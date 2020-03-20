from openpyxl import Workbook
from openpyxl.styles import Font
import glob
import os
from datetime import datetime
import math

os.chdir(r'D:\\PROJECT_RENAME\\')

wb = Workbook()

ws = wb.create_sheet(index =0,title ='ANU')
ws1 = wb.create_sheet(index =1,title ='Nava')
wb.save('nava.xlsx')

sheet=wb.get_sheet_by_name('ANU')

TITLE = ('FILE_NAME','DATE','SIZE')

sheet.append(TITLE)

os.chdir('D:\\PROJECT_RENAME')
print(os.getcwd())
F_Find = '*20200208*'
File =glob.glob(F_Find)
print(File)
LCOUNT= len(File)
print(LCOUNT)
for f in File:
    t = os.path.getmtime(f)
    tt =datetime.fromtimestamp(t).strftime('%Y:%m:%d')
    print(tt)
    update =f'{f},{tt}'
    sp = update.split(',')
    print(sp)
    sheet.append(sp)
    wb.save('nava.xlsx')
